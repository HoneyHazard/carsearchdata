#!/bin/python3
import os
import subprocess
import csv
import openpyxl as px
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile
from datetime import datetime
from bs4 import BeautifulSoup
import requests

def open_url_in_chromium(url):
    subprocess.Popen(['chromium-browser', url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

def get_input_from_user(prompt, default=""):
    user_input = input(f"{prompt} [{default}]: ").strip()
    return user_input if user_input else default

def get_hours_input(default=""):
    with NamedTemporaryFile(suffix="_hours.txt") as temp_file:
        subprocess.call(['nano', temp_file.name])
        with open(temp_file.name, 'r') as f:
            hours = f.read().strip().replace('\n', ', ')
    return hours if hours else default
    
def get_location_input(default=""):
    with NamedTemporaryFile(suffix="_address.txt") as temp_file:
        subprocess.call(['nano', temp_file.name])
        with open(temp_file.name, 'r') as f:
            location = f.read().strip().replace('\n', ', ')
    return location if location else default

def analyze_webpage(url):
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extracting details from Craigslist listing
        default_values = {}
        
        # link is also website
        default_values["website"] = url
        
        # Extract title
        title_element = soup.find("span", id="titletextonly")
        if title_element:
            default_values["year_model"] = title_element.text.strip()
        
        # Extract price
        price_element = soup.find("span", class_="price")
        if price_element:
            default_values["price"] = price_element.text.strip().replace("$", "")
            
        # Extract Location
        placename_meta = soup.find("meta", attrs={"name": "geo.placename"})
        region_meta = soup.find("meta", attrs={"name": "geo.region"})
        if placename_meta and region_meta:
            placename = placename_meta["content"]
            region = region_meta["content"]
            location = f"{placename}, {region.split('-')[-1]}"
            default_values["location"] = location
        else:
            default_values["location"] = ""

	# Extract Latitude & Longitude 
        geo_position_meta = soup.find("meta", attrs={"name": "geo.position"})
        if geo_position_meta:
            content = geo_position_meta["content"]
            latitude, longitude = content.split(";")
            default_values["latitude"] = latitude.strip()
            default_values["longitude"] = longitude.strip()
        else:
            default_values["latitude"] = ""
            default_values["longitude"] = ""
            
        # Extract posted date and generate status
        time_element = soup.find("time", class_="date timeago")
        if time_element:
            datetime_str = time_element["datetime"]
            posted_date = datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M:%S%z")
            default_values["status"] = f"posted {posted_date.strftime('%m/%d')}"
        else:
            default_values["status"] = ""
                       
        # Mapping between input fields and other Craigslist attr values
        mapping = {
            "odometer": "mileage",
            "condition": "condition",
            "cylinders": "engine",
            "drive": "drive",
            "title status": "accidents"
        }    
        attr_elements = soup.find_all("div", class_="attr")
        for attr_element in attr_elements:
            label_element = attr_element.find("span", class_="labl")
            value_element = attr_element.find("span", class_="valu")
            if label_element and value_element:
                label = label_element.text.strip().rstrip(":")
                value = value_element.text.strip()
                mapped_label = mapping.get(label, label)
                default_values[mapped_label] = value
        
        # Extract description
	#         description_element = soup.find("section", id="postingbody")
        # if description_element:
        #     default_values["Description"] = description_element.text.strip()
        
        return default_values
    except Exception as e:
        print(f"An error occurred while analyzing the Craigslist page: {e}")
        return {}

def write_to_excel(data, filename):
    if os.path.exists(filename):
        wb = px.load_workbook(filename)
        ws = wb.active
    else:
        wb = px.Workbook()
        ws = wb.active
        ws.append(['Car', 'Mileage', 'Price', 'Link', 'Drive', 'Engine', 'VIN', 'MPG', 'Accidents/Damage/Title', 
                   'Condition', 'Company', 'Website', 'Location', 'Latitude', 'Longitude', 'Hours', 'Contact', 'Phone', 'Email', 'Status'])
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
    
    ws.append(data)
    wb.save(filename)

def main():
    urls_file_or_url = input("Enter URLs file or a URL: ")

    if os.path.isfile(urls_file_or_url):
        with open(urls_file_or_url, 'r') as f:
            urls = f.readlines()
    else:
        urls = [urls_file_or_url.strip()]

    for url in urls:
        url = url.strip()
        open_url_in_chromium(url)
        
        default_values = analyze_webpage(url)
        
        car_info = []
        car_info.append(get_input_from_user("Year and model", default=default_values.get("year_model")))
        car_info.append(get_input_from_user("Mileage", default=default_values.get("mileage")))
        car_info.append(get_input_from_user("Price",  default=default_values.get("price")))
        car_info.append(f'=HYPERLINK("{url}"')
        car_info.append(get_input_from_user("Drive", default=default_values.get("drive")))
        car_info.append(get_input_from_user("Engine", default=default_values.get("engine")))
        car_info.append(get_input_from_user("VIN", default=default_values.get("vin")))
        car_info.append(get_input_from_user("MPG", default=default_values.get("mpg")))
        car_info.append(get_input_from_user("Accidents/Damage/Title", default=default_values.get("accidents")))
        car_info.append(get_input_from_user("Condition", default=default_values.get("condition")))
        car_info.append(get_input_from_user("Company/Person", default=default_values.get("company")))
        website = get_input_from_user("Website", default=default_values.get("website"))        
        car_info.append(f'=HYPERLINK("{website}")')
        # car_info.append(get_location_input(default=default_values.get("location")))
        car_info.append(get_input_from_user("Location", default=default_values.get("location")))
        car_info.append(default_values.get("latitude"))
        car_info.append(default_values.get("longitude"))
        # car_info.append(get_hours_input(default=default_values.get("hours")))
        car_info.append(get_input_from_user("Hours", default=default_values.get("hours")))
        car_info.append(get_input_from_user("Contact", default=default_values.get("contact")))
        car_info.append(get_input_from_user("Phone", default=default_values.get("phone")))
        car_info.append(get_input_from_user("Email", default=default_values.get("email")))
        car_info.append(get_input_from_user("Status", default=default_values.get("status")))

        write_to_excel(car_info, 'output.xlsx')
        
        # Displaying collected data before writing to Excel
        print("\n*** Collected Car Information: ***\n\n")
        for label, value in zip(['Year and model', 'Mileage', 'Price', 'Link', 'Drive', 'Engine', 'VIN', 'MPG', 'Accidents/Damage/Title', 'Condition', 'Company', 'Website', 'Location', 'Latitude', 'Longitude', 'Hours', 'Contact', 'Phone', 'Email', 'Status'], car_info):
            print(f"{label}: {value}")
        print("\n\n*** Next car... ***\n\n")

if __name__ == "__main__":
    main()

